# Write this in the terminal
# pip install requests python-dotenv openpyxl

# Imports
from dotenv import load_dotenv
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
import os
import requests
import json
import time
import openpyxl

# Load environment variables from .env file
load_dotenv()

api_key = os.getenv("API_KEY")
if not api_key:
    print("La clé API n'a pas été trouvée.")
else:
    # URL de base de l'API Riot
    base_url = "https://europe.api.riotgames.com"

    # Endpoint spécifique pour obtenir les informations du compte
    endpoint = "/riot/account/v1/accounts/by-riot-id/Colfeo/EUW"

    # Construire l'URL complète avec la clé API
    api_url_summoner = f"{base_url}{endpoint}?api_key={api_key}"

    try:
        # Effectuer la requête GET
        response = requests.get(api_url_summoner)

        # Vérifier si la requête a réussi
        if response.status_code == 200:
            player_info = response.json()

            # Afficher les informations du joueur de manière lisible
            #print(json.dumps(player_info, indent=4))

            # On veut maintenant récupérer les informations de la partie grâce au puuid
            puuid = player_info["puuid"]

            # Calculer les timestamps pour les dates spécifiées
            start_date = datetime(2025, 4, 24, 23, 59, 59)
            end_date = datetime(2025, 4, 27, 23, 59, 59)
            start_time = int(time.mktime(start_date.timetuple()))
            end_time = int(time.mktime(end_date.timetuple()))
            count = int(25)

            # Endpoint pour obtenir les informations de la partie
            endpoint = f"/lol/match/v5/matches/by-puuid/{puuid}/ids"
            # Construire l'URL complète avec la clé API et les paramètres de temps
            api_url_match = f"{base_url}{endpoint}?count={count}&startTime={start_time}&endTime={end_time}&api_key={api_key}"

            # Effectuer la requête GET
            response = requests.get(api_url_match)

            # Vérifier si la requête a réussi
            if response.status_code == 200:
                match_info = response.json()
                # Afficher les informations de la partie de manière lisible
                #print(json.dumps(match_info, indent=4))

                # Créer un nouveau fichier Excel
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Game et stats"

                # Définir les en-têtes de colonnes
                headers = ["Pseudo", "Champion", "KDA", "Kills", "Deaths", "Assists", "Cs", "Cs/m", "Dmg", "KP", "Vision score"]
                ws.append(headers)

                # Styles pour les en-têtes
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center')
                for col_num, column_title in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.alignment = header_alignment

                # On veut maintenant récupérer les informations de chaque partie, et enregistrer les données dans le fichier Excel
                for match_id in match_info:
                    # Il y a une partie buguée, on l'enlève
                    if(match_id == "EUW1_7381728303"):
                        continue

                    # Endpoint pour obtenir les détails de la partie
                    endpoint = f"/lol/match/v5/matches/{match_id}"
                    api_url_match_details = f"{base_url}{endpoint}?api_key={api_key}"

                    # Effectuer la requête GET
                    response = requests.get(api_url_match_details)

                    # Vérifier si la requête a réussi
                    if response.status_code == 200:
                        match_details = response.json()

                        # Extraire les informations nécessaires
                        game_duration = match_details["info"]["gameDuration"]
                        participants = match_details["info"]["participants"]

                        # Récupération le nombre total de kills de l'équipe et 
                        totalGameKills = 0                        
                        for participant in participants:                            
                            summoner_name = participant["riotIdGameName"]
                            if summoner_name in ["Colfeo", "climber", "Tha\u00efs Morel", "ThiBee", "Xintox"]:
                                totalGameKills += participant["kills"]
                            if summoner_name == "Colfeo":
                                # On veut savoir si on a gagné ou perdu
                                if participant["win"]:
                                    match_result = "Win"
                                else:
                                    match_result = "Lose"
                            if summoner_name == "Cygola":
                                opponent = "EC Emerald"
                            if summoner_name == "Killycurly":
                                opponent = "Mickey Mouse Club"
                            if summoner_name == "Guts":
                                opponent = "Burger Frites"
                            if summoner_name == "Tayme":
                                opponent = "Pipi gaming"
                            if summoner_name == "Zaliarell":
                                opponent = "LIONS GUARD"
                            if summoner_name == "Lasagna":
                                opponent = "PCS Geniuses"
                            
                        # On commence par mettre l'id de la partie, win ou lose et la durée de la game (minutes:secondes)            
                        head_data = [match_id, opponent, match_result, f"{game_duration // 60}:{game_duration % 60:02d}"] 
                        ws.append(head_data)

                        # Colorer la ligne en fonction du résultat
                        header_row = ws.max_row
                        for i in range(1, len(head_data) + 1):
                            cell = ws.cell(row=header_row, column=i)
                            if match_result == "Win":
                                cell.fill = openpyxl.styles.PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Vert clair
                            else:
                                cell.fill = openpyxl.styles.PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Rouge clair

                        # Le reste
                        for participant in participants:
                            summoner_name = participant["riotIdGameName"]
                            if summoner_name in ["Colfeo", "climber", "Tha\u00efs Morel", "ThiBee", "Xintox"]:             
                                champion = participant["championName"]
                                if(champion == "MonkeyKing"):
                                    champion = "Wukong"
                                #kda = (kills+assists)/deaths
                                kda = "PERFECT" if participant["deaths"] == 0 else f"{(participant['kills'] + participant['assists']) / participant['deaths']:.2f}"
                                kills = participant["kills"]
                                deaths = participant["deaths"]
                                assists = participant["assists"]
                                cs = participant["totalMinionsKilled"] + participant["neutralMinionsKilled"]
                                cs_per_min = cs / (game_duration / 60)
                                damage = participant["totalDamageDealtToChampions"]
                                #kp = (kills + assists) / all team kill (percentage)
                                kp = ((kills + assists) / totalGameKills) * 100 if totalGameKills > 0 else 0
                                vision_score = participant["visionScore"]

                                if(summoner_name == "climber"):
                                    summoner_name = "Mad"
                                if(summoner_name == "Tha\u00efs Morel"):
                                    summoner_name = "Sacha"
                                if(summoner_name == "ThiBee"):
                                    summoner_name = "Thib"

                                # Ajouter les données au fichier Excel
                                row_data = [summoner_name, champion, kda, kills, deaths, assists, cs, cs_per_min, damage, kp, vision_score]
                                ws.append(row_data)                                
                    else:
                        print(f"Erreur lors de la requête : {response.status_code}")
                        print("Réponse :", response.text)                   
                    # Ajouter une ligne séparatrice noire
                    separator_row = ws.max_row + 1
                    ws.append([])
                    for i in range(1, len(headers) + 1):
                        cell = ws.cell(row=separator_row, column=i)
                        cell.fill = openpyxl.styles.PatternFill(start_color="000000", end_color="000000", fill_type="solid")

                # Ajuster la largeur des colonnes
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Figer la première ligne
                ws.freeze_panes = 'A2'

                # Créer une feuille de résumé
                ws_summary = wb.create_sheet(title="Résumé")

                # Variables pour stocker les statistiques
                total_games = 0
                wins = 0
                losses = 0
                fastest_game = {"duration": float('inf'), "opponent": "", "match_id": ""}
                longest_game = {"duration": 0, "opponent": "", "match_id": ""}
                total_duration = 0

                # Dictionnaire pour stocker les statistiques par joueur
                player_stats = {
                    "Colfeo": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "damage": 0, "vision": 0, "kp": 0},
                    "Mad": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "damage": 0, "vision": 0, "kp": 0},
                    "Sacha": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "damage": 0, "vision": 0, "kp": 0},
                    "Thib": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "damage": 0, "vision": 0, "kp": 0},
                    "Xintox": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "damage": 0, "vision": 0, "kp": 0}
                }

                # Parcourir toutes les lignes du tableau de stats pour extraire les données
                row_index = 1  # Ignorer l'en-tête
                while row_index < ws.max_row:
                    row_index += 1
                    
                    # Si c'est une ligne d'en-tête de match
                    if ws.cell(row=row_index, column=1).value and ws.cell(row=row_index, column=1).value.startswith("EUW1"):
                        match_id = ws.cell(row=row_index, column=1).value
                        opponent = ws.cell(row=row_index, column=2).value
                        result = ws.cell(row=row_index, column=3).value
                        duration_str = ws.cell(row=row_index, column=4).value
                        
                        # Convertir la durée en secondes
                        minutes, seconds = map(int, duration_str.split(':'))
                        duration_seconds = minutes * 60 + seconds
                        
                        # Mettre à jour les statistiques globales
                        total_games += 1
                        total_duration += duration_seconds
                        
                        if result == "Win":
                            wins += 1
                        else:
                            losses += 1
                            
                        # Vérifier si c'est la partie la plus rapide ou la plus longue
                        if duration_seconds < fastest_game["duration"]:
                            fastest_game = {"duration": duration_seconds, "opponent": opponent, "match_id": match_id}
                        if duration_seconds > longest_game["duration"]:
                            longest_game = {"duration": duration_seconds, "opponent": opponent, "match_id": match_id}
                        
                        # Avancer pour lire les statistiques des joueurs
                        for i in range(5):  # Il y a 5 joueurs dans l'équipe
                            player_row = row_index + i + 1
                            if player_row < ws.max_row and ws.cell(row=player_row, column=1).value in player_stats:
                                player_name = ws.cell(row=player_row, column=1).value
                                player_stats[player_name]["games"] += 1
                                player_stats[player_name]["kills"] += ws.cell(row=player_row, column=4).value
                                player_stats[player_name]["deaths"] += ws.cell(row=player_row, column=5).value
                                player_stats[player_name]["assists"] += ws.cell(row=player_row, column=6).value
                                player_stats[player_name]["cs"] += ws.cell(row=player_row, column=7).value
                                player_stats[player_name]["damage"] += ws.cell(row=player_row, column=9).value
                                player_stats[player_name]["vision"] += ws.cell(row=player_row, column=11).value
                                player_stats[player_name]["kp"] += ws.cell(row=player_row, column=10).value
                    
                    # Sauter les lignes de données et séparateurs
                    while row_index < ws.max_row and not (ws.cell(row=row_index + 1, column=1).value and isinstance(ws.cell(row=row_index + 1, column=1).value, str) and ws.cell(row=row_index + 1, column=1).value.startswith("EUW1")):
                        row_index += 1

                # Calculer les moyennes et autres statistiques dérivées
                avg_duration = total_duration / total_games if total_games > 0 else 0
                winrate = (wins / total_games * 100) if total_games > 0 else 0

                # Ajouter les statistiques globales à la feuille de résumé
                ws_summary.append(["Statistiques globales de l'équipe"])
                ws_summary.append(["Nombre de parties jouées", total_games])
                ws_summary.append(["Nombre de victoires", wins])
                ws_summary.append(["Nombre de défaites", losses])
                ws_summary.append(["Winrate", f"{winrate:.2f}%"])
                ws_summary.append(["Durée moyenne des parties", f"{int(avg_duration // 60)}:{int(avg_duration % 60):02d}"])
                ws_summary.append(["Partie la plus rapide", f"{int(fastest_game['duration'] // 60)}:{int(fastest_game['duration'] % 60):02d}", f"contre {fastest_game['opponent']}", fastest_game['match_id']])
                ws_summary.append(["Partie la plus longue", f"{int(longest_game['duration'] // 60)}:{int(longest_game['duration'] % 60):02d}", f"contre {longest_game['opponent']}", longest_game['match_id']])
                ws_summary.append([])  # Ligne vide

                # Ajouter les statistiques par joueur
                ws_summary.append(["Statistiques par joueur"])
                ws_summary.append(["Joueur", "KDA", "Kills moy.", "Deaths moy.", "Assists moy.", "CS moy.", "Dmg moy.", "Vision moy.", "KP moy."])

                for player, stats in player_stats.items():
                    if stats["games"] > 0:
                        avg_kills = stats["kills"] / stats["games"]
                        avg_deaths = stats["deaths"] / stats["games"]
                        avg_assists = stats["assists"] / stats["games"]
                        # Format KDA identique à celui de la feuille "Game et stats"
                        kda = "PERFECT" if avg_deaths == 0 else f"{(avg_kills + avg_assists) / avg_deaths:.2f}"
                        avg_cs = stats["cs"] / stats["games"]
                        avg_damage = stats["damage"] / stats["games"]
                        avg_vision = stats["vision"] / stats["games"]
                        avg_kp = stats["kp"] / stats["games"]
                        
                        ws_summary.append([
                            player, 
                            kda,  # KDA au format numérique
                            f"{avg_kills:.2f}", 
                            f"{avg_deaths:.2f}", 
                            f"{avg_assists:.2f}", 
                            f"{avg_cs:.1f}", 
                            f"{avg_damage:.0f}", 
                            f"{avg_vision:.1f}", 
                            f"{avg_kp:.1f}%"
                        ])

                # Mettre en forme la feuille de résumé
                for row in ws_summary.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='left')

                # Ajuster la largeur des colonnes dans la feuille de résumé
                for column in ws_summary.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws_summary.column_dimensions[column_letter].width = adjusted_width

                # Mettre en évidence les titres
                title_font = Font(bold=True, size=14)
                ws_summary.cell(row=1, column=1).font = title_font
                ws_summary.cell(row=10, column=1).font = title_font

                # Mettre en forme les en-têtes des tableaux
                header_font = Font(bold=True)
                for col in range(1, 10):  # Ajusté pour tenir compte de la suppression de la colonne "Parties"
                    ws_summary.cell(row=11, column=col).font = header_font

                # Enregistrer le fichier Excel
                wb.save("Occi'lan #6.xlsx")
                print("Fichier Excel généré avec succès.")
            else:
                print(f"Erreur lors de la requête : {response.status_code}")
                print("Réponse :", response.text)
        else:
            print(f"Erreur lors de la requête : {response.status_code}")
            print("Réponse :", response.text)
    except requests.exceptions.RequestException as e:
        print(f"Une erreur s'est produite : {e}")