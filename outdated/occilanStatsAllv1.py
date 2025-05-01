# Write this in the terminal
# pip install requests python-dotenv openpyxl

# Imports
from dotenv import load_dotenv
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import requests
import json
import time
import openpyxl
import locale

# Définir la locale pour avoir des virgules au lieu des points pour les décimales
locale.setlocale(locale.LC_NUMERIC, 'fr_FR.UTF-8')

# Création des équipes via le fichier excel
def load_teams_from_excel(file_path):
    # Charger le fichier Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # Utiliser la première feuille

    teams = {}
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):  # Inclure les 4 colonnes
        team_name, role, player_name, player_tag = row
        if team_name not in teams:
            teams[team_name] = []
        teams[team_name].append({
            "role": role,
            "player_name": player_name,
            "player_tag": player_tag
        })

    return teams

# Exemple d'utilisation
current_dir = os.path.dirname(__file__)  # Récupère le dossier contenant le script Python
file_path = os.path.join(current_dir, "AllPseudosv2.xlsx")  # Construit le chemin vers le fichier Excel

teams = load_teams_from_excel(file_path)
print(json.dumps(teams, indent=4, ensure_ascii=False))  # `ensure_ascii=False` pour les caractères spéciaux

# Enregistrer les équipes dans un fichier JSON
output_json_path = os.path.join(current_dir, "teams.json")  # Chemin pour le fichier JSON
with open(output_json_path, "w", encoding="utf-8") as json_file:
    json.dump(teams, json_file, indent=4, ensure_ascii=False)  # `ensure_ascii=False` pour conserver les caractères spéciaux
    print(f"Les équipes ont été enregistrées dans {output_json_path}")

# Charger les variables d'environnement
load_dotenv()
api_key = os.getenv("API_KEY")
base_url = "https://europe.api.riotgames.com"

# Charger les équipes depuis le fichier JSON
current_dir = os.path.dirname(__file__)
teams_json_path = os.path.join(current_dir, "teams.json")

with open(teams_json_path, "r", encoding="utf-8") as json_file:
    teams = json.load(json_file)

# Ajouter les puuid pour chaque joueur
for team_name, players in teams.items():
    for player in players:
        player_name = player["player_name"]
        player_tag = player["player_tag"]

        # Endpoint pour récupérer le puuid
        endpoint = f"/riot/account/v1/accounts/by-riot-id/{player_name}/{player_tag}"
        api_url = f"{base_url}{endpoint}?api_key={api_key}"

        try:
            response = requests.get(api_url)
            if response.status_code == 200:
                player["puuid"] = response.json()["puuid"]
                print(f"puuid récupéré pour {player_name}#{player_tag}: {player['puuid']}")
            else:
                print(f"Erreur pour {player_name}#{player_tag}: {response.status_code} - {response.text}")
        except requests.exceptions.RequestException as e:
            print(f"Erreur lors de la requête pour {player_name}#{player_tag}: {e}")

# Sauvegarder les équipes avec les puuid dans un nouveau fichier JSON
output_json_path = os.path.join(current_dir, "teams_with_puuid.json")
with open(output_json_path, "w", encoding="utf-8") as json_file:
    json.dump(teams, json_file, indent=4, ensure_ascii=False)
    print(f"Les équipes avec les puuid ont été enregistrées dans {output_json_path}")