import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import os

def format_duration(minutes):
    """Convertit les minutes en format MM:SS"""
    total_minutes = int(minutes)
    seconds = int((minutes - total_minutes) * 60)
    return f"{total_minutes}:{seconds:02d}"

def create_tournament_sheet(wb, team_stats, general_stats):
    """Crée la feuille avec les statistiques globales du tournoi"""
    ws = wb.create_sheet("Tournoi")
    
    # Style des titres
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Titre principal
    ws['A1'] = "Statistiques globales du tournoi"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    
    # En-têtes
    headers = ["Statistique", "Joueur", "Équipe", "Valeur"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Stats globales
    row = 3
    for stat, data in general_stats.items():
        ws.cell(row=row, column=1, value=stat.replace('_', ' ').capitalize())
        ws.cell(row=row, column=2, value=data.get("player", ""))
        ws.cell(row=row, column=3, value=data.get("team", ""))
        ws.cell(row=row, column=4, value=data.get("value", ""))
        
        # Centrer les cellules
        for col in range(1, 5):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        row += 1

    # Ajuster les colonnes
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25

def create_team_sheet(wb, team_name, team_data, match_details):
    """Crée une feuille pour une équipe spécifique"""
    ws = wb.create_sheet(team_name[:31])  # Excel limite les noms d'onglets à 31 caractères
    
    # Styles
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Section 1: Statistiques globales de l'équipe
    ws['A1'] = f"Statistiques globales de l'équipe {team_name}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')

    row = 3
    team_stats = team_data["team_stats"]
    global_stats = [
        ("Nombre de parties jouées", team_stats["total_games"]),
        ("Nombre de victoires", team_stats["wins"]),
        ("Nombre de défaites", team_stats["losses"]),
        ("Winrate", f"{team_stats['winrate']}%"),
        ("Durée moyenne des parties", format_duration(team_stats["average_game_duration"])),
        ("Partie la plus rapide", format_duration(team_stats["shortest_game"])),
        ("Partie la plus longue", format_duration(team_stats["longest_game"]))
    ]

    for stat, value in global_stats:
        ws.cell(row=row, column=1, value=stat)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Section 2: Statistiques par joueur
    row += 2
    ws.cell(row=row, column=1, value="Statistiques par joueur")
    ws.cell(row=row, column=1).font = title_font
    row += 1

    # En-têtes des stats joueurs
    headers = ["Joueur", "KDA", "Kills moy.", "Deaths moy.", "Assists moy.", 
              "CS/M moy.", "Vision moy.", "Champions joués"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row += 1
    for player_name, player_stats in team_data["players"].items():
        ws.cell(row=row, column=1, value=player_name)
        ws.cell(row=row, column=2, value=player_stats["average_kda"])
        ws.cell(row=row, column=3, value=player_stats["average_kills"])
        ws.cell(row=row, column=4, value=player_stats["average_deaths"])
        ws.cell(row=row, column=5, value=player_stats["average_assists"])
        ws.cell(row=row, column=6, value=player_stats["average_cs_per_minute"])
        ws.cell(row=row, column=7, value=player_stats["average_vision_score"])
        ws.cell(row=row, column=8, value=", ".join(player_stats["champions_played"]))
        
        # Centrer les cellules sauf la dernière colonne (champions)
        for col in range(1, 8):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        row += 1

    # Section 3: Détail des parties
    row += 2
    ws.cell(row=row, column=1, value="Détail des parties")
    ws.cell(row=row, column=1).font = title_font
    row += 1

    # En-têtes des matchs
    match_headers = ["Match ID", "Joueur", "Champion", "KDA", "K/D/A", 
                    "CS", "CS/min", "Vision Score", "Résultat", "Durée"]
    
    for col, header in enumerate(match_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Ajuster les colonnes
    for col in range(1, len(match_headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_excel(team_stats, general_stats, match_details, output_path):
    """Crée le fichier Excel complet"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut
    
    # Créer la feuille du tournoi
    create_tournament_sheet(wb, team_stats, general_stats)
    
    # Créer une feuille par équipe
    for team_name, team_data in team_stats.items():
        create_team_sheet(wb, team_name, team_data, match_details)
    
    wb.save(output_path)
    print(f"Fichier Excel créé avec succès : {output_path}")

if __name__ == "__main__":
    current_dir = os.path.dirname(__file__)
    data_dir = os.path.join(current_dir, "../data")
    output_dir = os.path.join(current_dir, "../output")
    
    os.makedirs(output_dir, exist_ok=True)
    
    # Charger les données
    team_stats_path = os.path.join(data_dir, "team_stats.json")
    general_stats_path = os.path.join(data_dir, "general_stats.json")
    match_details_path = os.path.join(data_dir, "match_details.json")
    excel_output_path = os.path.join(output_dir, "OccilanStats.xlsx")
    
    with open(team_stats_path, "r", encoding="utf-8") as f:
        team_stats = json.load(f)
    with open(general_stats_path, "r", encoding="utf-8") as f:
        general_stats = json.load(f)
    with open(match_details_path, "r", encoding="utf-8") as f:
        match_details = json.load(f)
    
    create_excel(team_stats, general_stats, match_details, excel_output_path)