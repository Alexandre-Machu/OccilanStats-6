# Write this in the terminal
# pip install requests python-dotenv openpyxl

# Imports
from dotenv import load_dotenv
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
import requests
import json
import time
import openpyxl
import locale

# Définir la locale pour avoir des virgules au lieu des points pour les décimales
locale.setlocale(locale.LC_NUMERIC, 'fr_FR.UTF-8')

# Load environment variables from .env file
load_dotenv()

api_key = os.getenv("API_KEY")
if not api_key:
    print("La clé API n'a pas été trouvée.")
else:
    # URL de base de l'API Riot
    base_url = "https://europe.api.riotgames.com"

    # Endpoint spécifique pour obtenir les informations du compte
    endpoint = "/riot/account/v1/accounts/by-riot-id/Colfeo/EUW"

    # Construire l'URL complète avec la clé API
    api_url_summoner = f"{base_url}{endpoint}?api_key={api_key}"

    try:
        # Effectuer la requête GET
        response = requests.get(api_url_summoner)

        # Vérifier si la requête a réussi
        if response.status_code == 200:
            player_info = response.json()

            # On veut maintenant récupérer les informations de la partie grâce au puuid
            puuid = player_info["puuid"]

            # Calculer les timestamps pour les dates spécifiées
            start_date = datetime(2025, 4, 23)
            end_date = datetime(2025, 4, 27, 23, 59, 59)
            start_time = int(time.mktime(start_date.timetuple()))
            end_time = int(time.mktime(end_date.timetuple()))
            count = int(50)
            type = "tourney"

            # Endpoint pour obtenir les informations de la partie
            endpoint = f"/lol/match/v5/matches/by-puuid/{puuid}/ids"
            # Construire l'URL complète avec la clé API et les paramètres de temps
            api_url_match = f"{base_url}{endpoint}?count={count}&type={type}&startTime={start_time}&endTime={end_time}&api_key={api_key}"

            # Effectuer la requête GET
            response = requests.get(api_url_match)

            # Vérifier si la requête a réussi
            if response.status_code == 200:
                match_info = response.json()

                # Créer un nouveau fichier Excel
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Game et stats"

                # Définir les en-têtes de colonnes
                headers = ["Pseudo", "Champion", "KDA", "Kills", "Deaths", "Assists", "Cs", "Cs/m", "Dmg", "KP", "Vision score"]
                ws.append(headers)

                # Styles pour les en-têtes
                header_font = Font(bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center')
                header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                for col_num, column_title in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill
                    cell.border = thin_border

                # On veut maintenant récupérer les informations de chaque partie, et enregistrer les données dans le fichier Excel
                for match_id in match_info:
                    # Il y a une partie buguée, on l'enlève
                    if(match_id == "EUW1_7381728303"):
                        continue

                    # Endpoint pour obtenir les détails de la partie
                    endpoint = f"/lol/match/v5/matches/{match_id}"
                    api_url_match_details = f"{base_url}{endpoint}?api_key={api_key}"

                    # Effectuer la requête GET
                    response = requests.get(api_url_match_details)

                    # Vérifier si la requête a réussi
                    if response.status_code == 200:
                        match_details = response.json()

                        # Extraire les informations nécessaires
                        game_duration = match_details["info"]["gameDuration"]
                        participants = match_details["info"]["participants"]

                        # Récupération le nombre total de kills de l'équipe
                        totalGameKills = 0                        
                        for participant in participants:                            
                            summoner_name = participant["riotIdGameName"]
                            if summoner_name in ["Colfeo", "climber", "Tha\u00efs Morel", "ThiBee", "Xintox"]:
                                totalGameKills += participant["kills"]
                            if summoner_name == "Colfeo":
                                # On veut savoir si on a gagné ou perdu
                                if participant["win"]:
                                    match_result = "Win"
                                else:
                                    match_result = "Lose"
                            if summoner_name == "Cygola":
                                opponent = "EC Emerald"
                            if summoner_name == "Killycurly":
                                opponent = "Mickey Mouse Club"
                            if summoner_name == "Guts":
                                opponent = "Burger Frites"
                            if summoner_name == "Tayme":
                                opponent = "Pipi gaming"
                            if summoner_name == "Zaliarell":
                                opponent = "LIONS GUARD"
                            if summoner_name == "Lasagna":
                                opponent = "PCS Geniuses"
                            
                        # On commence par mettre l'id de la partie, win ou lose et la durée de la game (minutes:secondes)            
                        head_data = [match_id, opponent, match_result, f"{game_duration // 60}:{game_duration % 60:02d}"] 
                        ws.append(head_data)

                        # Colorer la ligne en fonction du résultat et appliquer un style
                        header_row = ws.max_row
                        header_font = Font(bold=True, size=11)
                        
                        for i in range(1, len(head_data) + 1):
                            cell = ws.cell(row=header_row, column=i)
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border
                            
                            if match_result == "Win":
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair

                        # Le reste
                        for participant in participants:
                            summoner_name = participant["riotIdGameName"]
                            if summoner_name in ["Colfeo", "climber", "Tha\u00efs Morel", "ThiBee", "Xintox"]:             
                                champion = participant["championName"]
                                if(champion == "MonkeyKing"):
                                    champion = "Wukong"
                                
                                #kda = (kills+assists)/deaths
                                kills = participant["kills"]
                                deaths = participant["deaths"]
                                assists = participant["assists"]
                                kda = "PERFECT" if deaths == 0 else f"{(kills + assists) / deaths:.2f}".replace(".", ",")
                                
                                cs = participant["totalMinionsKilled"] + participant["neutralMinionsKilled"]
                                cs_per_min = cs / (game_duration / 60)
                                cs_per_min_str = f"{cs_per_min:.1f}".replace(".", ",")
                                
                                damage = participant["totalDamageDealtToChampions"]
                                
                                #kp = (kills + assists) / all team kill (percentage)
                                kp = ((kills + assists) / totalGameKills) * 100 if totalGameKills > 0 else 0
                                kp_str = f"{kp:.1f}%".replace(".", ",")
                                
                                vision_score = participant["visionScore"]

                                if(summoner_name == "climber"):
                                    summoner_name = "Mad"
                                if(summoner_name == "Tha\u00efs Morel"):
                                    summoner_name = "Sacha"
                                if(summoner_name == "ThiBee"):
                                    summoner_name = "Thib"

                                # Ajouter les données au fichier Excel
                                row_data = [summoner_name, champion, kda, kills, deaths, assists, cs, cs_per_min_str, damage, kp_str, vision_score]
                                ws.append(row_data)
                                
                                # Appliquer le style aux cellules
                                player_row = ws.max_row
                                for col_num, value in enumerate(row_data, 1):
                                    cell = ws.cell(row=player_row, column=col_num)
                                    cell.alignment = Alignment(horizontal='center')
                                    
                                    # Pour les colonnes numériques, s'assurer qu'elles sont au format numérique
                                    if col_num in [4, 5, 6, 7, 9, 11]:  # Kills, Deaths, Assists, Cs, Dmg, Vision score
                                        if isinstance(value, str) and value.replace(".", "").isdigit():
                                            cell.value = float(value.replace(",", "."))
                                        else:
                                            cell.value = value
                                    else:
                                        cell.value = value
                                    
                                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))
                                
                                # Ajouter des bordures spéciales aux lignes spécifiées (E2:K2, E9:K9, E16:K16, etc.)
                                if player_row % 7 == 2:  # Vérifie si c'est une ligne qui nécessite une bordure spéciale
                                    for col_num in range(5, 12):  # Colonnes E à K (5 à 11 en index)
                                        cell = ws.cell(row=player_row, column=col_num)
                                        cell.border = thin_border
                    else:
                        print(f"Erreur lors de la requête : {response.status_code}")
                        print("Réponse :", response.text)                   
                    
                    # Ajouter une ligne séparatrice noire
                    separator_row = ws.max_row + 1
                    ws.append([])
                    for i in range(1, len(headers) + 1):
                        cell = ws.cell(row=separator_row, column=i)
                        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

                # Créer une feuille de résumé
                ws_summary = wb.create_sheet(title="Résumé")

                # Variables pour stocker les statistiques
                total_games = 0
                wins = 0
                losses = 0
                fastest_game = {"duration": float('inf'), "opponent": "", "match_id": ""}
                longest_game = {"duration": 0, "opponent": "", "match_id": ""}
                total_duration = 0

                # Dictionnaire pour stocker les statistiques par joueur
                player_stats = {
                    "Colfeo": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "cs_per_min": 0, "damage": 0, "vision": 0, "kp": 0},
                    "Mad": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "cs_per_min": 0, "damage": 0, "vision": 0, "kp": 0},
                    "Sacha": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "cs_per_min": 0, "damage": 0, "vision": 0, "kp": 0},
                    "Thib": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "cs_per_min": 0, "damage": 0, "vision": 0, "kp": 0},
                    "Xintox": {"games": 0, "kills": 0, "deaths": 0, "assists": 0, "cs": 0, "cs_per_min": 0, "damage": 0, "vision": 0, "kp": 0}
                }

                # Parcourir toutes les lignes du tableau de stats pour extraire les données
                row_index = 1  # Ignorer l'en-tête
                while row_index < ws.max_row:
                    row_index += 1
                    
                    # Si c'est une ligne d'en-tête de match
                    if ws.cell(row=row_index, column=1).value and ws.cell(row=row_index, column=1).value.startswith("EUW1"):
                        match_id = ws.cell(row=row_index, column=1).value
                        opponent = ws.cell(row=row_index, column=2).value
                        result = ws.cell(row=row_index, column=3).value
                        duration_str = ws.cell(row=row_index, column=4).value
                        
                        # Convertir la durée en secondes
                        minutes, seconds = map(int, duration_str.split(':'))
                        duration_seconds = minutes * 60 + seconds
                        
                        # Mettre à jour les statistiques globales
                        total_games += 1
                        total_duration += duration_seconds
                        
                        if result == "Win":
                            wins += 1
                        else:
                            losses += 1
                            
                        # Vérifier si c'est la partie la plus rapide ou la plus longue
                        if duration_seconds < fastest_game["duration"]:
                            fastest_game = {"duration": duration_seconds, "opponent": opponent, "match_id": match_id}
                        if duration_seconds > longest_game["duration"]:
                            longest_game = {"duration": duration_seconds, "opponent": opponent, "match_id": match_id}
                        
                        # Avancer pour lire les statistiques des joueurs
                        for i in range(5):  # Il y a 5 joueurs dans l'équipe
                            player_row = row_index + i + 1
                            if player_row < ws.max_row and ws.cell(row=player_row, column=1).value in player_stats:
                                player_name = ws.cell(row=player_row, column=1).value
                                player_stats[player_name]["games"] += 1
                                
                                # Récupérer les valeurs numériques en s'assurant qu'elles sont bien des nombres
                                kills_val = ws.cell(row=player_row, column=4).value
                                player_stats[player_name]["kills"] += float(kills_val) if isinstance(kills_val, str) else kills_val
                                
                                deaths_val = ws.cell(row=player_row, column=5).value
                                player_stats[player_name]["deaths"] += float(deaths_val) if isinstance(deaths_val, str) else deaths_val
                                
                                assists_val = ws.cell(row=player_row, column=6).value
                                player_stats[player_name]["assists"] += float(assists_val) if isinstance(assists_val, str) else assists_val
                                
                                cs_val = ws.cell(row=player_row, column=7).value
                                player_stats[player_name]["cs"] += float(cs_val) if isinstance(cs_val, str) else cs_val
                                
                                # Récupérer le CS/M en le convertissant correctement
                                cs_per_min_str = ws.cell(row=player_row, column=8).value
                                if isinstance(cs_per_min_str, str):
                                    cs_per_min_str = cs_per_min_str.replace(",", ".")
                                player_stats[player_name]["cs_per_min"] += float(cs_per_min_str)
                                
                                dmg_val = ws.cell(row=player_row, column=9).value
                                player_stats[player_name]["damage"] += float(dmg_val) if isinstance(dmg_val, str) else dmg_val
                                
                                # Récupérer le KP en le convertissant correctement
                                kp_str = ws.cell(row=player_row, column=10).value
                                if isinstance(kp_str, str):
                                    kp_str = kp_str.replace("%", "").replace(",", ".")
                                player_stats[player_name]["kp"] += float(kp_str)
                                
                                vision_val = ws.cell(row=player_row, column=11).value
                                player_stats[player_name]["vision"] += float(vision_val) if isinstance(vision_val, str) else vision_val
                    
                    # Sauter les lignes de données et séparateurs
                    while row_index < ws.max_row and not (ws.cell(row=row_index + 1, column=1).value and isinstance(ws.cell(row=row_index + 1, column=1).value, str) and ws.cell(row=row_index + 1, column=1).value.startswith("EUW1")):
                        row_index += 1

                # Calculer les moyennes et autres statistiques dérivées
                avg_duration = total_duration / total_games if total_games > 0 else 0
                winrate = (wins / total_games * 100) if total_games > 0 else 0

                # Styles pour la feuille de résumé
                title_font = Font(bold=True, size=14)
                subtitle_font = Font(bold=True, size=12)
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
                cell_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Ajouter les statistiques globales à la feuille de résumé
                ws_summary.append(["Statistiques globales de l'équipe"])
                ws_summary.cell(row=1, column=1).font = title_font
                
                global_stats = [
                    ["Nombre de parties jouées", total_games],
                    ["Nombre de victoires", wins],
                    ["Nombre de défaites", losses],
                    ["Winrate", f"{winrate:.2f}%".replace(".", ",")],
                    ["Durée moyenne des parties", f"{int(avg_duration // 60)}:{int(avg_duration % 60):02d}"],
                    ["Partie la plus rapide", f"{int(fastest_game['duration'] // 60)}:{int(fastest_game['duration'] % 60):02d}", f"contre {fastest_game['opponent']}", fastest_game['match_id']],
                    ["Partie la plus longue", f"{int(longest_game['duration'] // 60)}:{int(longest_game['duration'] % 60):02d}", f"contre {longest_game['opponent']}", longest_game['match_id']]
                ]
                
                for row_data in global_stats:
                    ws_summary.append(row_data)
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws_summary.cell(row=ws_summary.max_row, column=col_num)
                        cell.alignment = Alignment(horizontal='left')
                        cell.border = cell_border
                
                ws_summary.append([])  # Ligne vide

                # Ajouter les statistiques par joueur
                ws_summary.append(["Statistiques par joueur"])
                ws_summary.cell(row=ws_summary.max_row, column=1).font = title_font
                
                player_headers = ["Joueur", "KDA", "Kills moy.", "Deaths moy.", "Assists moy.", "CS/M moy.", "Dmg moy.", "Vision moy.", "KP moy."]
                ws_summary.append(player_headers)
                
                # Formater les en-têtes
                for col_num, value in enumerate(player_headers, 1):
                    cell = ws_summary.cell(row=ws_summary.max_row, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = cell_border

                for player, stats in player_stats.items():
                    if stats["games"] > 0:
                        avg_kills = stats["kills"] / stats["games"]
                        avg_deaths = stats["deaths"] / stats["games"]
                        avg_assists = stats["assists"] / stats["games"]
                        kda = "PERFECT" if avg_deaths == 0 else f"{(avg_kills + avg_assists) / avg_deaths:.2f}".replace(".", ",")
                        avg_cs_per_min = stats["cs_per_min"] / stats["games"]
                        avg_damage = stats["damage"] / stats["games"]
                        avg_vision = stats["vision"] / stats["games"]
                        avg_kp = stats["kp"] / stats["games"]
                        
                        player_row = [
                            player, 
                            kda,
                            f"{avg_kills:.2f}".replace(".", ","), 
                            f"{avg_deaths:.2f}".replace(".", ","), 
                            f"{avg_assists:.2f}".replace(".", ","), 
                            f"{avg_cs_per_min:.1f}".replace(".", ","), 
                            f"{avg_damage:.0f}".replace(".0", ""), 
                            f"{avg_vision:.1f}".replace(".", ","), 
                            f"{avg_kp:.1f}%".replace(".", ",")
                        ]
                        
                        ws_summary.append(player_row)
                        
                        # Formater les cellules de données
                        for col_num, value in enumerate(player_row, 1):
                            cell = ws_summary.cell(row=ws_summary.max_row, column=col_num)
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = cell_border

                # Ajuster la largeur des colonnes dans les deux feuilles
                for sheet in [ws, ws_summary]:
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        sheet.column_dimensions[column_letter].width = adjusted_width

                # Figer la première ligne
                ws.freeze_panes = 'A2'
                ws_summary.freeze_panes = 'A2'

                # Enregistrer le fichier Excel
                wb.save("Occi'lan #6 - La Ruche.xlsx")
                print("Fichier Excel généré avec succès.")
            else:
                print(f"Erreur lors de la requête : {response.status_code}")
                print("Réponse :", response.text)
        else:
            print(f"Erreur lors de la requête : {response.status_code}")
            print("Réponse :", response.text)
    except requests.exceptions.RequestException as e:
        print(f"Une erreur s'est produite : {e}")