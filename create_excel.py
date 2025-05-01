import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import json
import os

def calculate_additional_stats(match_details):
    """
    Calcule les statistiques supplémentaires à partir des détails des matchs.
    """
    longest_game = {"detail": None, "value": 0}
    shortest_game = {"detail": None, "value": float("inf")}
    most_kills_in_game = {"detail": None, "value": 0}
    least_kills_in_game = {"detail": None, "value": float("inf")}
    champion_picks = {}
    champion_bans = {}

    # Parcourir les matchs
    for match_id, match_data in match_details.items():
        try:
            # Assurez-vous que gameDuration est un nombre
            game_duration = float(match_data["info"].get("gameDuration", 0)) / 60  # Durée en minutes
            participants = match_data["info"].get("participants", [])
            teams = match_data["info"].get("teams", [])

            # Calculer la durée des parties
            if game_duration > longest_game["value"]:
                longest_game = {"detail": f"{match_id}", "value": f"{int(game_duration)}:{int((game_duration % 1) * 60):02d}"}
            if game_duration < shortest_game["value"]:
                shortest_game = {"detail": f"{match_id}", "value": f"{int(game_duration)}:{int((game_duration % 1) * 60):02d}"}

            # Calculer les kills dans la partie
            total_kills = sum(int(p.get("kills", 0)) for p in participants)  # Convertir les kills en entier
            if total_kills > most_kills_in_game["value"]:
                most_kills_in_game = {"detail": f"{match_id}", "value": total_kills}
            if total_kills < least_kills_in_game["value"]:
                least_kills_in_game = {"detail": f"{match_id}", "value": total_kills}

            # Compter les champions joués
            for participant in participants:
                champion = participant.get("championName")
                if champion:
                    champion_picks[champion] = champion_picks.get(champion, 0) + 1

            # Compter les champions bannis
            for team in teams:
                for ban in team.get("bans", []):
                    champion_id = ban.get("championId")
                    if champion_id:
                        champion_bans[champion_id] = champion_bans.get(champion_id, 0) + 1

        except KeyError as e:
            print(f"Clé manquante dans les données du match {match_id}: {e}")
        except ValueError as e:
            print(f"Erreur de conversion dans les données du match {match_id}: {e}")
        except Exception as e:
            print(f"Erreur inattendue lors du traitement du match {match_id}: {e}")

    # Trouver les champions les plus et les moins sélectionnés
    most_selected_champion = max(champion_picks.items(), key=lambda x: x[1], default=("Aucun", 0))
    least_selected_champion = min(champion_picks.items(), key=lambda x: x[1], default=("Aucun", 0))

    # Trouver les champions les plus et les moins bannis
    most_banned_champion = max(champion_bans.items(), key=lambda x: x[1], default=("Aucun", 0))
    least_banned_champion = min(champion_bans.items(), key=lambda x: x[1], default=("Aucun", 0))

    # Retourner les statistiques supplémentaires
    return {
        "longest_game": longest_game,
        "shortest_game": shortest_game,
        "most_kills_in_game": most_kills_in_game,
        "least_kills_in_game": least_kills_in_game,
        "most_selected_champion": {"detail": most_selected_champion[0], "value": most_selected_champion[1]},
        "least_selected_champion": {"detail": least_selected_champion[0], "value": least_selected_champion[1]},
        "most_banned_champion": {"detail": most_banned_champion[0], "value": most_banned_champion[1]},
        "least_banned_champion": {"detail": least_banned_champion[0], "value": least_banned_champion[1]}
    }

def create_tournament_sheet(wb, most_stats, additional_stats):
    """
    Crée la feuille "Tournoi" avec les statistiques globales.
    """
    sheet = wb.active
    sheet.title = "Tournoi"

    # Titre
    sheet.merge_cells("A1:E1")
    sheet["A1"] = "Résumé des statistiques globales du tournoi"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].alignment = Alignment(horizontal="center")

    # Ajouter les statistiques globales
    sheet.append(["Statistique", "Joueur/Partie/Champion", "Équipe", "Valeur"])

    # Parcourir les statistiques globales
    for stat, data in most_stats.items():
        sheet.append([
            stat.replace("_", " ").capitalize(),  # Remplacer les underscores par des espaces
            data["player"],
            data["team"],
            data["value"]
        ])

    # Ajouter les statistiques supplémentaires
    sheet.append([])  # Ligne vide pour séparer les sections
    sheet.append(["Statistique", "Détail", "Valeur"])

    for stat, data in additional_stats.items():
        sheet.append([
            stat.replace("_", " ").capitalize(),
            data["detail"],
            data["value"]
        ])

    # Ajuster la largeur des colonnes
    for col in range(1, 5):
        sheet.column_dimensions[get_column_letter(col)].width = 30

if __name__ == "__main__":
    # Charger les fichiers nécessaires
    current_dir = os.path.dirname(__file__)
    most_stats_path = os.path.join(current_dir, "most_stats.json")
    match_details_path = os.path.join(current_dir, "match_details.json")
    excel_output_path = os.path.join(current_dir, "OccilanStats.xlsx")

    # Charger les statistiques globales depuis most_stats.json
    with open(most_stats_path, "r", encoding="utf-8") as json_file:
        most_stats = json.load(json_file)

    # Charger les détails des matchs depuis match_details.json
    with open(match_details_path, "r", encoding="utf-8") as json_file:
        match_details = json.load(json_file)

    # Calculer les statistiques supplémentaires
    additional_stats = calculate_additional_stats(match_details)

    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    create_tournament_sheet(wb, most_stats, additional_stats)

    # Sauvegarder le fichier Excel
    wb.save(excel_output_path)
    print(f"Le fichier Excel a été enregistré dans {excel_output_path}")